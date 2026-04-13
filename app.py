from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime, date
from dateutil.relativedelta import relativedelta

app = Flask(__name__, static_folder='static', template_folder='templates')
CORS(app)
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024

EXCEL_FILE     = 'financas_casa.xlsx'
UPLOAD_DIR     = 'comprovantes'
ALLOWED_EXT    = {'png', 'jpg', 'jpeg', 'webp', 'gif', 'pdf'}
PROJECTION_MONTHS = 24

SHEETS = {
    'transacoes': 'Transações',
    'compras':    'Compras Casa',
    'contas':     'Contas Fixas',
    'filho':      'Gastos Filho',
    'resumo':     'Resumo Mensal',
}

HEADER_FILL = PatternFill("solid", start_color="1A1A2E", end_color="1A1A2E")
HEADER_FONT = Font(bold=True, color="E94560", size=11)
FILHO_FONT  = Font(bold=True, color="06D6A0", size=11)
ALT_FILL    = PatternFill("solid", start_color="16213E", end_color="16213E")
NORMAL_FILL = PatternFill("solid", start_color="0F3460", end_color="0F3460")
NORMAL_FONT = Font(color="E2E8F0", size=10)
BORDER = Border(
    left=Side(style='thin', color='2D3748'), right=Side(style='thin', color='2D3748'),
    top=Side(style='thin', color='2D3748'), bottom=Side(style='thin', color='2D3748'),
)

def _setup_sheet(ws, headers, widths, accent=None):
    ws.sheet_view.showGridLines = False
    font = accent or HEADER_FONT
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = font; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 25
    ws.freeze_panes = 'A2'

def _style_row(ws, row_num, alt=False):
    fill = ALT_FILL if alt else NORMAL_FILL
    for cell in ws[row_num]:
        cell.fill = fill; cell.font = NORMAL_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

def get_next_id(ws):
    ids = [int(r[0].value) for r in ws.iter_rows(min_row=2)
           if r[0].value and str(r[0].value).isdigit()]
    return max(ids) + 1 if ids else 1

def _normalize_key(k):
    if not k: return k
    mapping = str.maketrans(
        'ÁÀÂÃáàâãÉÈÊéèêÍÌÎíìîÓÒÔÕóòôõÚÙÛúùûÇç',
        'AAAAaaaaEEEeeeIIIiiiOOOOooooUUUuuuCc')
    return str(k).translate(mapping)

def load_sheet(name):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[name]
    raw_headers = [c.value for c in ws[1]]
    headers = [_normalize_key(h) for h in raw_headers]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row): continue
        d = {}
        for h, v in zip(headers, row):
            if isinstance(v, (datetime, date)):
                d[h] = v.strftime('%Y-%m-%d') if isinstance(v, datetime) else str(v)
            else:
                d[h] = v
        data.append(d)
    return data

def allowed(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXT

def del_file(fname):
    if fname and isinstance(fname, str):
        p = os.path.join(UPLOAD_DIR, fname)
        if os.path.exists(p): os.remove(p)

def _add_month(dt_str, months):
    dt = datetime.strptime(str(dt_str)[:10], '%Y-%m-%d')
    return (dt + relativedelta(months=months)).strftime('%Y-%m-%d')

def _due_date(base_date_str, dia_vencimento, month_offset=0):
    import calendar
    base = datetime.strptime(str(base_date_str)[:10], '%Y-%m-%d')
    target = base + relativedelta(months=month_offset)
    last_day = calendar.monthrange(target.year, target.month)[1]
    day = min(int(dia_vencimento), last_day)
    return target.replace(day=day).strftime('%Y-%m-%d')

def _add_col(ws, col_name, width, default=None):
    hdrs = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if col_name in hdrs:
        return False
    nc = ws.max_column + 1
    cell = ws.cell(row=1, column=nc, value=col_name)
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER
    ws.column_dimensions[get_column_letter(nc)].width = width
    if default is not None:
        for row in ws.iter_rows(min_row=2):
            if any(c.value for c in row):
                ws.cell(row=row[0].row, column=nc, value=default)
    return True

COMPRAS_HEADERS = [
    'ID','Data Compra','Item','Categoria','Loja/Fornecedor','Responsavel',
    'Valor Total (R$)','Valor Parcela (R$)','Status','Prioridade',
    'Num Parcelas','Parcela Atual','Dia Vencimento','Data Vencimento',
    'Status Pagamento','Observacao','Comprovante'
]
COMPRAS_WIDTHS = [6,13,28,20,22,14,16,16,13,11,12,13,14,15,16,24,28]

FILHO_HEADERS = [
    'ID','Data','Descricao','Categoria','Responsavel',
    'Valor Total (R$)','Valor Parcela (R$)','Status',
    'Num Parcelas','Parcela Atual','Dia Vencimento','Data Vencimento',
    'Status Pagamento','Comprovante','Observacao','Recorrente'
]
FILHO_WIDTHS = [8,12,30,25,15,16,16,12,12,12,14,15,15,30,25,12]

def init_excel():
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    if os.path.exists(EXCEL_FILE):
        _migrate(); return
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet(SHEETS['transacoes'])
    _setup_sheet(ws, ['ID','Data','Descricao','Categoria','Tipo','Responsavel','Valor (R$)','Observacao','Comprovante','Recorrente'], [8,12,30,22,12,15,15,25,30,12])
    ws = wb.create_sheet(SHEETS['compras'])
    _setup_sheet(ws, COMPRAS_HEADERS, COMPRAS_WIDTHS)
    ws = wb.create_sheet(SHEETS['contas'])
    _setup_sheet(ws, ['ID','Nome da Conta','Categoria','Valor (R$)','Dia Vencimento','Responsavel','Status','Mes Referencia','Observacao','Comprovante','Recorrente'], [8,30,20,15,15,15,12,15,25,30,12])
    ws = wb.create_sheet(SHEETS['filho'])
    _setup_sheet(ws, FILHO_HEADERS, FILHO_WIDTHS, accent=FILHO_FONT)
    ws = wb.create_sheet(SHEETS['resumo'])
    _setup_sheet(ws, ['Mes/Ano','Total Receitas (R$)','Total Despesas (R$)','Total Compras (R$)','Total Contas (R$)','Total Filho (R$)','Saldo (R$)','Status'], [15,20,20,20,18,18,15,12])
    wb.save(EXCEL_FILE)

def _migrate():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    changed = False
    if SHEETS['transacoes'] in wb.sheetnames:
        ws = wb[SHEETS['transacoes']]
        changed |= _add_col(ws, 'Comprovante', 30, '')
        changed |= _add_col(ws, 'Recorrente', 12, 'Não')
    if SHEETS['compras'] in wb.sheetnames:
        ws = wb[SHEETS['compras']]
        changed |= _add_col(ws, 'Comprovante', 28, '')
        changed |= _add_col(ws, 'Num Parcelas', 12, 1)
        changed |= _add_col(ws, 'Parcela Atual', 13, 1)
        changed |= _add_col(ws, 'Dia Vencimento', 14, 10)
        changed |= _add_col(ws, 'Data Vencimento', 15, '')
        changed |= _add_col(ws, 'Status Pagamento', 16, 'Pendente')
        changed |= _add_col(ws, 'Valor Total (R$)', 16, None)
        changed |= _add_col(ws, 'Valor Parcela (R$)', 16, None)
        hdrs2 = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

        def _cidx(name):
            return hdrs2.index(name) + 1 if name in hdrs2 else None

        col_data  = _cidx('Data') or _cidx('Data Compra')
        col_valor = _cidx('Valor (R$)')
        col_vt    = _cidx('Valor Total (R$)')
        col_vp    = _cidx('Valor Parcela (R$)')
        col_np    = _cidx('Num Parcelas')
        col_pa    = _cidx('Parcela Atual')
        col_dia   = _cidx('Dia Vencimento')
        col_dv    = _cidx('Data Vencimento')
        col_stpag = _cidx('Status Pagamento')
        col_st    = _cidx('Status')
        col_prio  = _cidx('Prioridade')
        col_obs   = _cidx('Observacao')
        PAGO_STATUS = {'Comprado', 'Pago'}

        for row in ws.iter_rows(min_row=2):
            if not any(c.value for c in row):
                continue
            changed = True

            st_raw = row[col_st - 1].value if col_st else None
            is_corrupt = False
            if st_raw is not None:
                try: float(st_raw); is_corrupt = True
                except (TypeError, ValueError): pass
            if is_corrupt:
                obs_real = row[col_stpag - 1].value if col_stpag else None
                dia_raw  = row[col_dia - 1].value if col_dia else '2026-03-23'
                try:
                    dia_int = datetime.strptime(str(dia_raw)[:10], '%Y-%m-%d').day
                except Exception:
                    try: dia_int = int(dia_raw or 10)
                    except: dia_int = 10
                pa_raw = row[col_pa - 1].value if col_pa else None
                try: pa_int = int(pa_raw or 1)
                except: pa_int = 1
                if pa_int > 31: pa_int = 1
                if col_st:    row[col_st - 1].value    = 'Parcelado'
                if col_prio:  row[col_prio - 1].value  = 'Alta'
                if col_obs:   row[col_obs - 1].value   = obs_real
                if col_np:    row[col_np - 1].value    = 1
                if col_pa:    row[col_pa - 1].value    = 1
                if col_dia:   row[col_dia - 1].value   = dia_int
                if col_stpag: row[col_stpag - 1].value = 'Pendente'
                if col_dv and col_data:
                    raw = row[col_data - 1].value
                    if raw:
                        base = raw.strftime('%Y-%m-%d') if isinstance(raw,(datetime,date)) else str(raw)[:10]
                        row[col_dv - 1].value = _due_date(base, dia_int, 0)
                continue

            if col_valor and col_vt and row[col_vt - 1].value is None:
                row[col_vt - 1].value = row[col_valor - 1].value or 0
            if col_vt and col_vp and col_np:
                vt = float(row[col_vt - 1].value or 0)
                np = int(row[col_np - 1].value or 1)
                if row[col_vp - 1].value is None:
                    row[col_vp - 1].value = round(vt / max(np, 1), 2)

            st_val = str(row[col_st - 1].value or '') if col_st else ''
            if col_stpag:
                cur_stpag = row[col_stpag - 1].value
                if cur_stpag is None:
                    row[col_stpag - 1].value = 'Pago' if st_val in PAGO_STATUS else 'Pendente'
                elif st_val in PAGO_STATUS and str(cur_stpag) not in ('Pago',):
                    row[col_stpag - 1].value = 'Pago'

            if col_dv and row[col_dv - 1].value in (None, ''):
                base_date = None
                if col_data:
                    raw = row[col_data - 1].value
                    if raw:
                        if isinstance(raw, (datetime, date)):
                            base_date = raw.strftime('%Y-%m-%d')
                        else:
                            base_date = str(raw)[:10]
                if base_date:
                    dia_raw = row[col_dia - 1].value if col_dia else None
                    try:
                        dia = int(dia_raw or 10)
                    except (ValueError, TypeError):
                        try:
                            dia = datetime.strptime(str(dia_raw)[:10], '%Y-%m-%d').day
                            if col_dia: row[col_dia - 1].value = dia
                        except Exception:
                            dia = 10
                    pa  = int(row[col_pa - 1].value or 1) if col_pa else 1
                    row[col_dv - 1].value = _due_date(base_date, dia, pa - 1)

    if SHEETS['contas'] in wb.sheetnames:
        ws = wb[SHEETS['contas']]
        changed |= _add_col(ws, 'Comprovante', 30, '')
        changed |= _add_col(ws, 'Recorrente', 12, 'Sim')

    if SHEETS['filho'] not in wb.sheetnames:
        ws = wb.create_sheet(SHEETS['filho'])
        _setup_sheet(ws, FILHO_HEADERS, FILHO_WIDTHS, accent=FILHO_FONT)
        changed = True
    else:
        ws = wb[SHEETS['filho']]
        changed |= _add_col(ws, 'Valor Total (R$)', 16, None)
        changed |= _add_col(ws, 'Valor Parcela (R$)', 16, None)
        changed |= _add_col(ws, 'Status', 12, 'Pendente')
        changed |= _add_col(ws, 'Num Parcelas', 12, 1)
        changed |= _add_col(ws, 'Parcela Atual', 12, 1)
        changed |= _add_col(ws, 'Dia Vencimento', 14, 10)
        changed |= _add_col(ws, 'Data Vencimento', 15, None)
        changed |= _add_col(ws, 'Status Pagamento', 15, 'Pendente')
        changed |= _add_col(ws, 'Recorrente', 12, 'Não')

        hdrs_f = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        def _fidx(name): return hdrs_f.index(name) + 1 if name in hdrs_f else None

        col_fvr    = _fidx('Valor (R$)')
        col_fvt    = _fidx('Valor Total (R$)')
        col_fvp    = _fidx('Valor Parcela (R$)')
        col_fdt    = _fidx('Data')
        col_fdia   = _fidx('Dia Vencimento')
        col_fdv    = _fidx('Data Vencimento')
        col_fpa    = _fidx('Parcela Atual')
        col_fst    = _fidx('Status')
        col_fnp    = _fidx('Num Parcelas')
        col_fstpag = _fidx('Status Pagamento')
        col_fcomp  = _fidx('Comprovante')
        col_fobs   = _fidx('Observacao')

        for row in ws.iter_rows(min_row=2):
            if not any(c.value for c in row): continue
            changed = True

            comp_raw = row[col_fcomp - 1].value if col_fcomp else None
            is_corrupt_f = False
            if comp_raw is not None:
                try:
                    float(comp_raw)
                    is_corrupt_f = True
                except (ValueError, TypeError):
                    pass

            if is_corrupt_f:
                valor_total = float(row[col_fvr - 1].value or 0) if col_fvr else 0
                dia_raw_st = row[col_fst - 1].value if col_fst else 10
                try:
                    dia_fixado = int(dia_raw_st)
                except (ValueError, TypeError):
                    dia_fixado = 10
                base_raw = row[col_fdt - 1].value if col_fdt else None
                if isinstance(base_raw, (datetime, date)):
                    base_str = base_raw.strftime('%Y-%m-%d')
                elif base_raw:
                    base_str = str(base_raw)[:10]
                else:
                    base_str = datetime.now().strftime('%Y-%m-%d')
                if col_fcomp:  row[col_fcomp  - 1].value = ''
                if col_fobs:   row[col_fobs   - 1].value = ''
                if col_fvt:    row[col_fvt    - 1].value = valor_total
                if col_fvp:    row[col_fvp    - 1].value = valor_total
                if col_fvr:    row[col_fvr    - 1].value = valor_total
                if col_fst:    row[col_fst    - 1].value = 'Pendente'
                if col_fnp:    row[col_fnp    - 1].value = 1
                if col_fpa:    row[col_fpa    - 1].value = 1
                if col_fdia:   row[col_fdia   - 1].value = dia_fixado
                if col_fdv:    row[col_fdv    - 1].value = _due_date(base_str, dia_fixado, 0)
                if col_fstpag: row[col_fstpag - 1].value = 'Pendente'
                continue

            if col_fvr and col_fvt and row[col_fvt - 1].value is None:
                row[col_fvt - 1].value = row[col_fvr - 1].value or 0
            if col_fvt and col_fvp and row[col_fvp - 1].value is None:
                row[col_fvp - 1].value = row[col_fvt - 1].value or 0
            if col_fdv and row[col_fdv - 1].value in (None, ''):
                raw = row[col_fdt - 1].value if col_fdt else None
                if raw:
                    base_date = raw.strftime('%Y-%m-%d') if isinstance(raw, (datetime, date)) else str(raw)[:10]
                    dia_raw_f = row[col_fdia - 1].value if col_fdia else None
                    try:
                        dia = int(dia_raw_f or 10)
                    except (ValueError, TypeError):
                        try:
                            dia = datetime.strptime(str(dia_raw_f)[:10], '%Y-%m-%d').day
                            if col_fdia: row[col_fdia - 1].value = dia
                        except Exception:
                            dia = 10
                    try:
                        pa = int(row[col_fpa - 1].value or 1) if col_fpa else 1
                    except (ValueError, TypeError):
                        pa = 1
                    row[col_fdv - 1].value = _due_date(base_date, dia, pa - 1)

    if changed: wb.save(EXCEL_FILE)

@app.route("/")
def index(): return send_from_directory('templates', 'index.html')
@app.route('/static/<path:path>')
def statics(path): return send_from_directory('static', path)

@app.route('/api/upload', methods=['POST'])
def upload():
    try:
        if 'file' not in request.files: return jsonify({'success':False,'error':'Sem arquivo'}),400
        f = request.files['file']
        if not f.filename or not allowed(f.filename): return jsonify({'success':False,'error':'Arquivo invalido'}),400
        ext=f.filename.rsplit('.',1)[1].lower(); categoria=request.form.get('categoria','SemCategoria')
        descricao=request.form.get('descricao',f.filename.rsplit('.',1)[0]); tipo=request.form.get('tipo','Geral')
        ts=datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        import re,unicodedata
        def _slug(s):
            s=''.join(c for c in s if ord(c)<128 or unicodedata.category(c).startswith('L'))
            s=unicodedata.normalize('NFKD',s).encode('ascii','ignore').decode('ascii')
            s=re.sub(r'[^a-zA-Z0-9]+','_',s).strip('_'); return s[:35] or 'sem_nome'
        fname=f"{_slug(tipo)}_{_slug(categoria)}_{_slug(descricao)}_{ts}.{ext}"
        f.save(os.path.join(UPLOAD_DIR,fname)); return jsonify({'success':True,'filename':fname})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/comprovantes/<filename>')
def serve_comp(filename): return send_from_directory(UPLOAD_DIR,filename)

# ── TRANSAÇÕES ────────────────────────────────────────────────────────────────
@app.route('/api/transacoes',methods=['GET'])
def get_trans():
    try: return jsonify({'success':True,'data':load_sheet(SHEETS['transacoes'])})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/transacoes',methods=['POST'])
def add_trans():
    try:
        b=request.json; wb=openpyxl.load_workbook(EXCEL_FILE); ws=wb[SHEETS['transacoes']]
        nid=get_next_id(ws); rn=ws.max_row+1
        rec=b.get('recorrente','Não')
        if b.get('tipo','')!='Receita': rec='Não'
        ws.append([nid,b.get('data',datetime.now().strftime('%Y-%m-%d')),b.get('descricao',''),b.get('categoria',''),b.get('tipo',''),b.get('responsavel',''),float(b.get('valor',0)),b.get('observacao',''),b.get('comprovante',''),rec])
        _style_row(ws,rn,alt=(rn%2==0)); ws.cell(row=rn,column=7).number_format='R$ #,##0.00'
        wb.save(EXCEL_FILE); _update_resumo(); return jsonify({'success':True,'id':nid})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/transacoes/<int:iid>',methods=['DELETE'])
def del_trans(iid): return _del(SHEETS['transacoes'],iid,comp_col=9)

@app.route('/api/transacoes/<int:iid>',methods=['PUT'])
def edit_trans(iid):
    try:
        b=request.json; rec=b.get('recorrente','Não')
        if b.get('tipo','')!='Receita': rec='Não'
        wb=openpyxl.load_workbook(EXCEL_FILE); ws=wb[SHEETS['transacoes']]
        for row in ws.iter_rows(min_row=2):
            if row[0].value==iid:
                row[1].value=b.get('data',row[1].value); row[2].value=b.get('descricao',row[2].value)
                row[3].value=b.get('categoria',row[3].value); row[4].value=b.get('tipo',row[4].value)
                row[5].value=b.get('responsavel',row[5].value); row[6].value=float(b.get('valor',row[6].value))
                row[6].number_format='R$ #,##0.00'; row[7].value=b.get('observacao',row[7].value)
                if b.get('comprovante'): row[8].value=b['comprovante']
                ws.cell(row=row[0].row,column=10).value=rec; break
        wb.save(EXCEL_FILE); _update_resumo(); return jsonify({'success':True})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

# ── COMPRAS ───────────────────────────────────────────────────────────────────
@app.route('/api/compras',methods=['GET'])
def get_comp():
    try: return jsonify({'success':True,'data':load_sheet(SHEETS['compras'])})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/compras',methods=['POST'])
def add_comp():
    try:
        b=request.json; wb=openpyxl.load_workbook(EXCEL_FILE); ws=wb[SHEETS['compras']]
        hdrs=[ws.cell(1,c).value for c in range(1,ws.max_column+1)]
        def ci(n): return hdrs.index(n)+1 if n in hdrs else None
        status=b.get('status','Pendente')
        PAGO_STATUS = {'Comprado','Pago'}
        num_parcelas=int(b.get('num_parcelas',1)) if status=='Parcelado' else 1
        dia_venc=int(b.get('dia_vencimento',10))
        valor_total=float(b.get('valor',0))
        val_parcela=round(valor_total/num_parcelas,2)
        base_date=b.get('data',datetime.now().strftime('%Y-%m-%d'))
        nid=get_next_id(ws)
        ci_id=ci('ID'); ci_data=ci('Data') or ci('Data Compra'); ci_item=ci('Item')
        ci_cat=ci('Categoria'); ci_loja=ci('Loja/Fornecedor'); ci_resp=ci('Responsavel')
        ci_vr=ci('Valor (R$)'); ci_vt=ci('Valor Total (R$)'); ci_vp=ci('Valor Parcela (R$)')
        ci_st=ci('Status'); ci_prio=ci('Prioridade'); ci_np=ci('Num Parcelas')
        ci_pa=ci('Parcela Atual'); ci_dia=ci('Dia Vencimento'); ci_dv=ci('Data Vencimento')
        ci_stpag=ci('Status Pagamento'); ci_obs=ci('Observacao'); ci_comp=ci('Comprovante')
        for i in range(num_parcelas):
            data_venc=_due_date(base_date,dia_venc,i)
            st_pag='Pago' if status in PAGO_STATUS else 'Pendente'
            rn=ws.max_row+1; cur_nid=nid+i
            row_vals=[None]*len(hdrs)
            def sv(col_idx,val):
                if col_idx: row_vals[col_idx-1]=val
            sv(ci_id,cur_nid); sv(ci_data,base_date); sv(ci_item,b.get('item',''))
            sv(ci_cat,b.get('categoria','')); sv(ci_loja,b.get('loja',''))
            sv(ci_resp,b.get('responsavel','')); sv(ci_vr,valor_total)
            sv(ci_vt,valor_total); sv(ci_vp,val_parcela); sv(ci_st,status)
            sv(ci_prio,b.get('prioridade','Média')); sv(ci_np,num_parcelas)
            sv(ci_pa,i+1); sv(ci_dia,dia_venc); sv(ci_dv,data_venc)
            sv(ci_stpag,st_pag); sv(ci_obs,b.get('observacao',''))
            sv(ci_comp,b.get('comprovante','') if i==0 else '')
            ws.append(row_vals); _style_row(ws,rn,alt=(rn%2==0))
            if ci_vt: ws.cell(row=rn,column=ci_vt).number_format='R$ #,##0.00'
            if ci_vp: ws.cell(row=rn,column=ci_vp).number_format='R$ #,##0.00'
            if ci_vr: ws.cell(row=rn,column=ci_vr).number_format='R$ #,##0.00'
        wb.save(EXCEL_FILE); _update_resumo()
        return jsonify({'success':True,'id':nid,'parcelas_criadas':num_parcelas})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/compras/<int:iid>',methods=['DELETE'])
def del_comp(iid): return _del(SHEETS['compras'],iid,comp_col=17)

@app.route('/api/compras/<int:iid>',methods=['PUT'])
def edit_comp(iid):
    try:
        b=request.json; wb=openpyxl.load_workbook(EXCEL_FILE); ws=wb[SHEETS['compras']]
        hdrs=[ws.cell(1,c).value for c in range(1,ws.max_column+1)]
        def ci(n): return hdrs.index(n)+1 if n in hdrs else None
        ci_id=ci('ID'); ci_data=ci('Data') or ci('Data Compra'); ci_item=ci('Item')
        ci_cat=ci('Categoria'); ci_loja=ci('Loja/Fornecedor'); ci_resp=ci('Responsavel')
        ci_vr=ci('Valor (R$)'); ci_vt=ci('Valor Total (R$)'); ci_vp=ci('Valor Parcela (R$)')
        ci_st=ci('Status'); ci_prio=ci('Prioridade'); ci_np=ci('Num Parcelas')
        ci_pa=ci('Parcela Atual'); ci_dia=ci('Dia Vencimento'); ci_dv=ci('Data Vencimento')
        ci_stpag=ci('Status Pagamento'); ci_obs=ci('Observacao'); ci_comp=ci('Comprovante')
        PAGO_STATUS = {'Comprado','Pago'}
        def gv(row, col_idx, fallback=None):
            return row[col_idx-1].value if col_idx else fallback

        target_row = None
        for row in ws.iter_rows(min_row=2):
            if row[0].value == iid:
                target_row = row; break
        if not target_row:
            return jsonify({'success':False,'error':'ID não encontrado'}),404

        orig_item = gv(target_row, ci_item) or ''
        orig_data = str(gv(target_row, ci_data) or '')[:10]
        orig_vt   = float(gv(target_row, ci_vt) or gv(target_row, ci_vr) or 0)

        novo_status     = b.get('status', str(gv(target_row, ci_st) or 'Pendente'))
        novo_np         = int(b.get('num_parcelas', 1)) if novo_status == 'Parcelado' else 1
        novo_valor      = float(b.get('valor', orig_vt))
        novo_parcela    = round(novo_valor / max(novo_np, 1), 2)
        novo_dia        = int(b.get('dia_vencimento', gv(target_row, ci_dia) or 10))
        novo_data       = b.get('data', orig_data)
        novo_item       = b.get('item', orig_item)
        novo_cat        = b.get('categoria', gv(target_row, ci_cat) or '')
        novo_loja       = b.get('loja', gv(target_row, ci_loja) or '')
        novo_resp       = b.get('responsavel', gv(target_row, ci_resp) or '')
        novo_prio       = b.get('prioridade', gv(target_row, ci_prio) or 'Média')
        novo_obs        = b.get('observacao', gv(target_row, ci_obs) or '')
        novo_comp       = b.get('comprovante', gv(target_row, ci_comp) or '')

        group_rows = []
        for row in ws.iter_rows(min_row=2):
            if not any(c.value for c in row): continue
            r_item = str(gv(row, ci_item) or '')
            r_data = str(gv(row, ci_data) or '')[:10]
            r_vt   = float(gv(row, ci_vt) or gv(row, ci_vr) or 0)
            if r_item == orig_item and r_data == orig_data and abs(r_vt - orig_vt) < 0.01:
                group_rows.append(row[0].row)

        pagas_map = {}
        for rn in group_rows:
            pa_val  = ws.cell(rn, ci_pa).value if ci_pa else 1
            stpag_v = ws.cell(rn, ci_stpag).value if ci_stpag else 'Pendente'
            st_v    = ws.cell(rn, ci_st).value if ci_st else ''
            pa_int  = int(pa_val or 1)
            is_pago = (str(stpag_v) == 'Pago' or str(st_v) in PAGO_STATUS)
            pagas_map[pa_int] = 'Pago' if is_pago else 'Pendente'

        for rn in sorted(group_rows, reverse=True):
            ws.delete_rows(rn)

        nid = get_next_id(ws)
        st_pag_global = 'Pago' if novo_status in PAGO_STATUS else 'Pendente'
        for i in range(novo_np):
            data_venc = _due_date(novo_data, novo_dia, i)
            pa_num    = i + 1
            st_pag    = pagas_map.get(pa_num, st_pag_global)
            if novo_status in PAGO_STATUS: st_pag = 'Pago'

            rn = ws.max_row + 1
            row_vals = [None] * len(hdrs)
            def sv(col_idx, val):
                if col_idx: row_vals[col_idx-1] = val
            sv(ci_id, nid+i); sv(ci_data, novo_data); sv(ci_item, novo_item)
            sv(ci_cat, novo_cat); sv(ci_loja, novo_loja); sv(ci_resp, novo_resp)
            sv(ci_vr, novo_valor); sv(ci_vt, novo_valor); sv(ci_vp, novo_parcela)
            sv(ci_st, novo_status); sv(ci_prio, novo_prio); sv(ci_np, novo_np)
            sv(ci_pa, pa_num); sv(ci_dia, novo_dia); sv(ci_dv, data_venc)
            sv(ci_stpag, st_pag); sv(ci_obs, novo_obs)
            sv(ci_comp, novo_comp if i == 0 else '')
            ws.append(row_vals); _style_row(ws, rn, alt=(rn%2==0))
            for col in [ci_vr, ci_vt, ci_vp]:
                if col: ws.cell(row=rn, column=col).number_format='R$ #,##0.00'

        wb.save(EXCEL_FILE); _update_resumo()
        return jsonify({'success':True, 'parcelas_criadas': novo_np})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/compras/<int:iid>/status',methods=['PATCH'])
def upd_comp_status(iid):
    try:
        wb=openpyxl.load_workbook(EXCEL_FILE); ws=wb[SHEETS['compras']]
        for row in ws.iter_rows(min_row=2):
            if row[0].value==iid: row[8].value=request.json.get('status',row[8].value); break
        wb.save(EXCEL_FILE); return jsonify({'success':True})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/compras/<int:iid>/pagar',methods=['PATCH'])
def pagar_parcela(iid):
    try:
        wb=openpyxl.load_workbook(EXCEL_FILE); ws=wb[SHEETS['compras']]
        hdrs=[ws.cell(1,c).value for c in range(1,ws.max_column+1)]
        ci_stpag=hdrs.index('Status Pagamento')+1 if 'Status Pagamento' in hdrs else 16
        for row in ws.iter_rows(min_row=2):
            if row[0].value==iid:
                ws.cell(row=row[0].row,column=ci_stpag).value='Pago'
                ws.cell(row=row[0].row,column=ci_stpag).font=Font(color='00FF88',bold=True)
                break
        wb.save(EXCEL_FILE); _update_resumo(); return jsonify({'success':True})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

# ── CONTAS ────────────────────────────────────────────────────────────────────
@app.route('/api/contas',methods=['GET'])
def get_contas():
    try: return jsonify({'success':True,'data':load_sheet(SHEETS['contas'])})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/contas',methods=['POST'])
def add_conta():
    try:
        b=request.json; wb=openpyxl.load_workbook(EXCEL_FILE); ws=wb[SHEETS['contas']]
        nid=get_next_id(ws); rn=ws.max_row+1
        ws.append([nid,b.get('nome',''),b.get('categoria',''),float(b.get('valor',0)),int(b.get('dia_vencimento',1)),b.get('responsavel',''),b.get('status','Pendente'),b.get('mes_referencia',datetime.now().strftime('%m/%Y')),b.get('observacao',''),b.get('comprovante',''),b.get('recorrente','Sim')])
        _style_row(ws,rn,alt=(rn%2==0)); ws.cell(row=rn,column=4).number_format='R$ #,##0.00'
        wb.save(EXCEL_FILE); _update_resumo(); return jsonify({'success':True,'id':nid})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/contas/<int:iid>',methods=['DELETE'])
def del_conta(iid): return _del(SHEETS['contas'],iid,comp_col=10)

@app.route('/api/contas/<int:iid>',methods=['PUT'])
def edit_conta(iid):
    try:
        b=request.json; wb=openpyxl.load_workbook(EXCEL_FILE); ws=wb[SHEETS['contas']]
        for row in ws.iter_rows(min_row=2):
            if row[0].value==iid:
                row[1].value=b.get('nome',row[1].value); row[2].value=b.get('categoria',row[2].value)
                row[3].value=float(b.get('valor',row[3].value)); row[3].number_format='R$ #,##0.00'
                row[4].value=int(b.get('dia_vencimento',row[4].value)); row[5].value=b.get('responsavel',row[5].value)
                row[6].value=b.get('status',row[6].value); row[7].value=b.get('mes_referencia',row[7].value)
                row[8].value=b.get('observacao',row[8].value)
                if b.get('comprovante'): row[9].value=b['comprovante']
                ws.cell(row=row[0].row,column=11).value=b.get('recorrente','Sim'); break
        wb.save(EXCEL_FILE); _update_resumo(); return jsonify({'success':True})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/contas/<int:iid>/pagar',methods=['PATCH'])
def pagar_conta(iid):
    try:
        wb=openpyxl.load_workbook(EXCEL_FILE); ws=wb[SHEETS['contas']]
        for row in ws.iter_rows(min_row=2):
            if row[0].value==iid: row[6].value='Pago'; row[6].font=Font(color='00FF88',bold=True); break
        wb.save(EXCEL_FILE); _update_resumo(); return jsonify({'success':True})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

# ── FILHO ─────────────────────────────────────────────────────────────────────
@app.route('/api/filho',methods=['GET'])
def get_filho():
    try: return jsonify({'success':True,'data':load_sheet(SHEETS['filho'])})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/filho',methods=['POST'])
def add_filho():
    try:
        b=request.json; wb=openpyxl.load_workbook(EXCEL_FILE); ws=wb[SHEETS['filho']]
        hdrs=[ws.cell(1,c).value for c in range(1,ws.max_column+1)]
        def ci(n): return hdrs.index(n)+1 if n in hdrs else None
        status=b.get('status','Pendente')
        num_parcelas=int(b.get('num_parcelas',1)) if status=='Parcelado' else 1
        dia_venc=int(b.get('dia_vencimento',10))
        valor_total=float(b.get('valor',0))
        val_parcela=round(valor_total/num_parcelas,2)
        base_date=b.get('data',datetime.now().strftime('%Y-%m-%d'))
        nid=get_next_id(ws)
        ci_id=ci('ID'); ci_data=ci('Data'); ci_desc=ci('Descricao')
        ci_cat=ci('Categoria'); ci_resp=ci('Responsavel')
        ci_vr=ci('Valor (R$)'); ci_vt=ci('Valor Total (R$)'); ci_vp=ci('Valor Parcela (R$)')
        ci_st=ci('Status'); ci_np=ci('Num Parcelas'); ci_pa=ci('Parcela Atual')
        ci_dia=ci('Dia Vencimento'); ci_dv=ci('Data Vencimento')
        ci_stpag=ci('Status Pagamento'); ci_comp=ci('Comprovante'); ci_obs=ci('Observacao')
        ci_rec=ci('Recorrente')
        for i in range(num_parcelas):
            rn=ws.max_row+1
            data_venc=_due_date(base_date,dia_venc,i)
            row_vals=[None]*len(hdrs)
            def sv(col_idx, val):
                if col_idx: row_vals[col_idx-1]=val
            sv(ci_id,   nid+i)
            sv(ci_data, base_date)
            sv(ci_desc, b.get('descricao',''))
            sv(ci_cat,  b.get('categoria',''))
            sv(ci_resp, b.get('responsavel',''))
            sv(ci_vt,   valor_total)
            sv(ci_vp,   val_parcela)
            sv(ci_vr,   valor_total)
            sv(ci_st,   status)
            sv(ci_np,   num_parcelas)
            sv(ci_pa,   i+1)
            sv(ci_dia,  dia_venc)
            sv(ci_dv,   data_venc)
            sv(ci_stpag,'Pendente')
            sv(ci_comp, b.get('comprovante','') if i==0 else '')
            sv(ci_obs,  b.get('observacao',''))
            sv(ci_rec,  'Não' if status == 'Parcelado' else b.get('recorrente','Não'))
            ws.append(row_vals)
            _style_row(ws,rn,alt=(rn%2==0))
            for col in [ci_vt, ci_vp, ci_vr]:
                if col: ws.cell(row=rn,column=col).number_format='R$ #,##0.00'
        wb.save(EXCEL_FILE); _update_resumo()
        return jsonify({'success':True,'id':nid,'parcelas_criadas':num_parcelas})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/filho/<int:iid>',methods=['DELETE'])
def del_filho(iid):
    try:
        wb=openpyxl.load_workbook(EXCEL_FILE); ws=wb[SHEETS['filho']]
        hdrs=[ws.cell(1,c).value for c in range(1,ws.max_column+1)]
        comp_col = hdrs.index('Comprovante')+1 if 'Comprovante' in hdrs else 7
        for row in ws.iter_rows(min_row=2):
            if row[0].value==iid:
                del_file(ws.cell(row=row[0].row,column=comp_col).value)
                ws.delete_rows(row[0].row); break
        wb.save(EXCEL_FILE); _update_resumo(); return jsonify({'success':True})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/filho/<int:iid>',methods=['PUT'])
def edit_filho(iid):
    try:
        b=request.json; wb=openpyxl.load_workbook(EXCEL_FILE); ws=wb[SHEETS['filho']]
        hdrs=[ws.cell(1,c).value for c in range(1,ws.max_column+1)]
        def _cidx(n): return hdrs.index(n)+1 if n in hdrs else None
        uses_new = 'Valor Total (R$)' in hdrs
        for row in ws.iter_rows(min_row=2):
            if row[0].value==iid:
                row[1].value=b.get('data',row[1].value)
                row[2].value=b.get('descricao',row[2].value)
                row[3].value=b.get('categoria',row[3].value)
                row[4].value=b.get('responsavel',row[4].value)
                if uses_new:
                    ci_vt  = _cidx('Valor Total (R$)')
                    ci_vp  = _cidx('Valor Parcela (R$)')
                    ci_st  = _cidx('Status')
                    ci_np  = _cidx('Num Parcelas')
                    ci_dia = _cidx('Dia Vencimento')
                    ci_dv  = _cidx('Data Vencimento')
                    ci_com = _cidx('Comprovante')
                    ci_obs = _cidx('Observacao')
                    if ci_vt: row[ci_vt-1].value=float(b.get('valor',row[ci_vt-1].value or 0)); row[ci_vt-1].number_format='R$ #,##0.00'
                    if ci_vp:
                        np_val=int(row[_cidx('Num Parcelas')-1].value or 1) if _cidx('Num Parcelas') else 1
                        row[ci_vp-1].value=round(float(b.get('valor',row[ci_vt-1].value or 0))/np_val,2)
                        row[ci_vp-1].number_format='R$ #,##0.00'
                    if ci_st and b.get('status'): row[ci_st-1].value=b['status']
                    if ci_dia and b.get('dia_vencimento'):
                        row[ci_dia-1].value=int(b['dia_vencimento'])
                        if ci_dv: row[ci_dv-1].value=_due_date(str(row[1].value),int(b['dia_vencimento']))
                    if b.get('comprovante') and ci_com: row[ci_com-1].value=b['comprovante']
                    if ci_obs: row[ci_obs-1].value=b.get('observacao',row[ci_obs-1].value)
                    ci_rec2 = _cidx('Recorrente')
                    if ci_rec2:
                        st_atual = b.get('status', str(row[ci_st-1].value if ci_st else 'Pendente'))
                        row[ci_rec2-1].value = 'Não' if st_atual == 'Parcelado' else b.get('recorrente','Não')
                else:
                    row[5].value=float(b.get('valor',row[5].value)); row[5].number_format='R$ #,##0.00'
                    if b.get('comprovante'): row[6].value=b['comprovante']
                    row[7].value=b.get('observacao',row[7].value)
                break
        wb.save(EXCEL_FILE); _update_resumo(); return jsonify({'success':True})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/filho/<int:iid>/pagar',methods=['PATCH'])
def pagar_filho(iid):
    try:
        wb=openpyxl.load_workbook(EXCEL_FILE); ws=wb[SHEETS['filho']]
        hdrs=[ws.cell(1,c).value for c in range(1,ws.max_column+1)]
        ci_stpag = hdrs.index('Status Pagamento')+1 if 'Status Pagamento' in hdrs else None
        for row in ws.iter_rows(min_row=2):
            if row[0].value==iid:
                if ci_stpag:
                    ws.cell(row=row[0].row,column=ci_stpag).value='Pago'
                    ws.cell(row=row[0].row,column=ci_stpag).font=Font(color='00FF88',bold=True)
                break
        wb.save(EXCEL_FILE); _update_resumo(); return jsonify({'success':True})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

# ── PAGAMENTOS ────────────────────────────────────────────────────────────────
@app.route('/api/debug/compras')
def debug_compras():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb[SHEETS['compras']]
        headers_raw = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None for v in row): continue
            rows.append(dict(zip(headers_raw, [str(v) if v is not None else None for v in row])))
        return jsonify({'headers': headers_raw, 'rows': rows[:10]})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pagamentos',methods=['GET'])
def get_pagamentos():
    try:
        import calendar
        mes_filtro = request.args.get('mes', datetime.now().strftime('%Y-%m'))
        compras    = load_sheet(SHEETS['compras'])
        contas     = load_sheet(SHEETS['contas'])
        filho_data = load_sheet(SHEETS['filho'])
        today      = datetime.today()
        today_mes  = today.strftime('%Y-%m')   # ← usado para distinguir passado/presente de futuro
        horizon    = (today + relativedelta(months=PROJECTION_MONTHS)).strftime('%Y-%m-%d')
        items      = []

        def _safe_int_dia(raw, fallback=10):
            if raw is None: return fallback
            try: return int(raw)
            except (ValueError, TypeError):
                try: return datetime.strptime(str(raw)[:10], '%Y-%m-%d').day
                except: return fallback

        def _compute_dv(data_base, dia, pa_offset=0):
            if not data_base: return ''
            try:
                base = datetime.strptime(str(data_base)[:10], '%Y-%m-%d')
                target = base + relativedelta(months=pa_offset)
                ld = calendar.monthrange(target.year, target.month)[1]
                return target.replace(day=min(dia, ld)).strftime('%Y-%m-%d')
            except: return ''

        def _effective_dv(row_dict, data_keys, dia_key, pa_key):
            dv = str(row_dict.get('Data Vencimento') or '')[:10]
            if len(dv) == 10 and dv[:7]:
                return dv
            data_base = None
            for k in data_keys:
                v = row_dict.get(k)
                if v: data_base = str(v)[:10]; break
            dia = _safe_int_dia(row_dict.get(dia_key), 10)
            pa  = int(row_dict.get(pa_key) or 1)
            return _compute_dv(data_base, dia, pa - 1)

        # ── COMPRAS DA CASA ──────────────────────────────────────────────────
        comp_groups = {}
        for c in compras:
            key = (
                str(c.get('Item') or ''),
                str(c.get('Data') or c.get('Data Compra') or '')[:10],
                str(c.get('Valor Total (R$)') or c.get('Valor (R$)') or '0'),
            )
            comp_groups.setdefault(key, []).append(c)

        for key, grupo in comp_groups.items():
            grupo_sorted = sorted(grupo, key=lambda r: int(r.get('Parcela Atual') or 1))
            total_parc   = int(grupo_sorted[0].get('Num Parcelas') or len(grupo_sorted))
            pagas_grupo  = [r for r in grupo_sorted if str(r.get('Status Pagamento') or 'Pendente') == 'Pago']

            for c in grupo_sorted:
                dv = _effective_dv(c, ['Data', 'Data Compra'], 'Dia Vencimento', 'Parcela Atual')
                if not dv or dv[:7] != mes_filtro:
                    continue

                pa      = int(c.get('Parcela Atual') or 1)
                st_raw  = str(c.get('Status Pagamento') or 'Pendente')
                st_comp = str(c.get('Status') or '')
                st_pag  = 'Pago' if (st_raw == 'Pago' or st_comp in ('Comprado','Pago')) else st_raw
                val_p   = float(c.get('Valor Parcela (R$)') or c.get('Valor Total (R$)') or c.get('Valor (R$)') or 0)
                val_t   = float(c.get('Valor Total (R$)') or c.get('Valor (R$)') or 0)
                rest_tot= len([r for r in grupo_sorted if str(r.get('Status Pagamento') or 'Pendente') != 'Pago'])
                rest_dep= len([r for r in grupo_sorted
                               if str(r.get('Status Pagamento') or 'Pendente') != 'Pago'
                               and int(r.get('Parcela Atual') or 1) > pa])

                items.append({
                    'id':                  c.get('ID'),
                    'tipo':                'parcela',
                    'descricao':           c.get('Item') or '—',
                    'categoria':           c.get('Categoria') or '—',
                    'responsavel':         c.get('Responsavel') or '—',
                    'valor':               val_p,
                    'valor_total':         val_t,
                    'data_vencimento':     dv,
                    'num_parcelas':        total_parc,
                    'parcela_atual':       pa,
                    'parcelas_pagas':      len(pagas_grupo),
                    'parcelas_restantes':  rest_tot,
                    'restantes_apos_esta': rest_dep,
                    'status_pagamento':    st_pag,
                    'prioridade':          c.get('Prioridade') or '—',
                    'loja':                c.get('Loja/Fornecedor') or '—',
                    'comprovante':         c.get('Comprovante') or '',
                })

        # ── CONTAS FIXAS ─────────────────────────────────────────────────────
        for c in contas:
            def cval(*keys):
                for k in keys:
                    v = c.get(k)
                    if v is not None: return v
                return None

            rec    = str(cval('Recorrente') or 'Sim')
            mr     = str(cval('Mes Referencia') or '')
            dia    = _safe_int_dia(cval('Dia Vencimento'), 1)
            val    = float(cval('Valor (R$)') or 0)
            status = str(cval('Status') or 'Pendente')

            if '/' in mr:
                p = mr.split('/'); mes_base = f"{p[1]}-{p[0].zfill(2)}"
            else:
                mes_base = today.strftime('%Y-%m')

            relevant = [mes_base]
            if rec == 'Sim':
                try: cur = datetime.strptime(mes_base + '-01', '%Y-%m-%d')
                except: cur = today
                for _ in range(PROJECTION_MONTHS):
                    cur += relativedelta(months=1)
                    relevant.append(cur.strftime('%Y-%m'))
                    if cur.strftime('%Y-%m-%d') > horizon: break

            if mes_filtro not in relevant: continue
            yr, mo = int(mes_filtro[:4]), int(mes_filtro[5:7])
            ld = calendar.monthrange(yr, mo)[1]
            dv = f"{mes_filtro}-{str(min(dia, ld)).zfill(2)}"

            # ── FIX: para meses passados/presentes usa o status real do registro;
            #         só força Pendente em meses futuros (ainda não chegaram).
            st = status if mes_filtro <= today_mes else 'Pendente'

            items.append({
                'id':                  cval('ID'),
                'tipo':                'conta_fixa',
                'descricao':           cval('Nome da Conta') or '—',
                'categoria':           cval('Categoria') or '—',
                'responsavel':         cval('Responsavel') or '—',
                'valor':               val,
                'valor_total':         val,
                'data_vencimento':     dv,
                'num_parcelas':        None,
                'parcela_atual':       None,
                'parcelas_pagas':      None,
                'parcelas_restantes':  None,
                'restantes_apos_esta': None,
                'status_pagamento':    st,
                'prioridade':          '—',
                'loja':                '—',
                'comprovante':         cval('Comprovante') or '',
                'recorrente':          rec,
            })

        # ── GASTOS DO FILHO ──────────────────────────────────────────────────
        filho_rec  = [f for f in filho_data
                      if str(f.get('Recorrente') or 'Não') == 'Sim'
                      and int(f.get('Num Parcelas') or 1) <= 1]
        filho_norm = [f for f in filho_data if f not in filho_rec]

        # Filho recorrente — projeta como conta fixa
        for f in filho_rec:
            dia    = _safe_int_dia(f.get('Dia Vencimento'), 10)
            val_t  = float(f.get('Valor Total (R$)') or f.get('Valor (R$)') or 0)
            st_pag = str(f.get('Status Pagamento') or f.get('Status') or 'Pendente')
            data_b = str(f.get('Data') or '')[:10]
            mes_base = data_b[:7] if len(data_b) >= 7 else today.strftime('%Y-%m')

            relevant = [mes_base]
            try: cur = datetime.strptime(mes_base + '-01', '%Y-%m-%d')
            except: cur = today
            for _ in range(PROJECTION_MONTHS):
                cur += relativedelta(months=1)
                relevant.append(cur.strftime('%Y-%m'))
                if cur.strftime('%Y-%m-%d') > horizon: break

            if mes_filtro not in relevant: continue
            yr2, mo2 = int(mes_filtro[:4]), int(mes_filtro[5:7])
            ld2 = calendar.monthrange(yr2, mo2)[1]
            dv2 = f"{mes_filtro}-{str(min(dia, ld2)).zfill(2)}"

            # ── FIX: para meses passados/presentes usa o status real do registro;
            #         só força Pendente em meses futuros (ainda não chegaram).
            st2 = st_pag if mes_filtro <= today_mes else 'Pendente'

            items.append({
                'id':                  f.get('ID'),
                'tipo':                'filho',
                'descricao':           f.get('Descricao') or '—',
                'categoria':           f.get('Categoria') or '—',
                'responsavel':         f.get('Responsavel') or '—',
                'valor':               val_t,
                'valor_total':         val_t,
                'data_vencimento':     dv2,
                'num_parcelas':        None,
                'parcela_atual':       None,
                'parcelas_pagas':      None,
                'parcelas_restantes':  None,
                'restantes_apos_esta': None,
                'status_pagamento':    st2,
                'prioridade':          '—',
                'loja':                '—',
                'comprovante':         f.get('Comprovante') or '',
                'recorrente':          'Sim',
            })

        # Filho não-recorrente / parcelado — agrupa por produto
        filho_groups = {}
        for f in filho_norm:
            key = (
                str(f.get('Descricao') or ''),
                str(f.get('Data') or '')[:10],
                str(f.get('Valor Total (R$)') or f.get('Valor (R$)') or '0'),
            )
            filho_groups.setdefault(key, []).append(f)

        for key, fgrupo in filho_groups.items():
            fgrupo_sorted = sorted(fgrupo, key=lambda r: int(r.get('Parcela Atual') or 1))
            total_parc_f  = int(fgrupo_sorted[0].get('Num Parcelas') or 1)
            pagas_f_grupo = [r for r in fgrupo_sorted if str(r.get('Status Pagamento') or 'Pendente') == 'Pago']

            for f in fgrupo_sorted:
                dv = _effective_dv(f, ['Data'], 'Dia Vencimento', 'Parcela Atual')
                if not dv:
                    dv = str(f.get('Data') or '')[:10]
                if not dv or dv[:7] != mes_filtro:
                    continue

                pa      = int(f.get('Parcela Atual') or 1)
                st_pag  = str(f.get('Status Pagamento') or f.get('Status') or 'Pendente')
                val_p   = float(f.get('Valor Parcela (R$)') or f.get('Valor Total (R$)') or f.get('Valor (R$)') or 0)
                val_t   = float(f.get('Valor Total (R$)') or f.get('Valor (R$)') or 0)
                rest_tot= len([r for r in fgrupo_sorted if str(r.get('Status Pagamento') or 'Pendente') != 'Pago'])
                rest_dep= len([r for r in fgrupo_sorted
                               if str(r.get('Status Pagamento') or 'Pendente') != 'Pago'
                               and int(r.get('Parcela Atual') or 1) > pa])

                items.append({
                    'id':                  f.get('ID'),
                    'tipo':                'filho',
                    'descricao':           f.get('Descricao') or '—',
                    'categoria':           f.get('Categoria') or '—',
                    'responsavel':         f.get('Responsavel') or '—',
                    'valor':               val_p,
                    'valor_total':         val_t,
                    'data_vencimento':     dv,
                    'num_parcelas':        total_parc_f,
                    'parcela_atual':       pa,
                    'parcelas_pagas':      len(pagas_f_grupo),
                    'parcelas_restantes':  rest_tot,
                    'restantes_apos_esta': rest_dep,
                    'status_pagamento':    st_pag,
                    'prioridade':          '—',
                    'loja':                '—',
                    'comprovante':         f.get('Comprovante') or '',
                })

        items.sort(key=lambda x: x['data_vencimento'] or '')

        total_mes      = sum(i['valor'] for i in items)
        total_pago     = sum(i['valor'] for i in items if i['status_pagamento'] == 'Pago')
        total_pendente = sum(i['valor'] for i in items if i['status_pagamento'] == 'Pendente')
        total_atrasado = sum(i['valor'] for i in items
                            if i['status_pagamento'] == 'Pendente'
                            and (i['data_vencimento'] or '') < today.strftime('%Y-%m-%d'))
        cat_totais = {}
        for i in items:
            cat = i['categoria']; cat_totais[cat] = cat_totais.get(cat, 0) + i['valor']
        tipo_totais = {
            'Compras':      sum(i['valor'] for i in items if i['tipo'] == 'parcela'),
            'Contas Fixas': sum(i['valor'] for i in items if i['tipo'] == 'conta_fixa'),
            'Filho':        sum(i['valor'] for i in items if i['tipo'] == 'filho'),
        }
        return jsonify({
            'success': True, 'mes': mes_filtro, 'items': items,
            'resumo': {'total': total_mes, 'pago': total_pago, 'pendente': total_pendente, 'atrasado': total_atrasado},
            'por_categoria': cat_totais, 'por_tipo': tipo_totais,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ── DASHBOARD ─────────────────────────────────────────────────────────────────
@app.route('/api/dashboard',methods=['GET'])
def dashboard():
    try:
        import calendar as cal_mod
        trans=load_sheet(SHEETS['transacoes']); compras=load_sheet(SHEETS['compras'])
        contas=load_sheet(SHEETS['contas']); filho=load_sheet(SHEETS['filho'])
        rec=sum(float(t.get('Valor (R$)',0) or 0) for t in trans if t.get('Tipo')=='Receita')
        des=sum(float(t.get('Valor (R$)',0) or 0) for t in trans if t.get('Tipo')=='Despesa')
        tc=sum(float(c.get('Valor Parcela (R$)') or c.get('Valor Total (R$)') or c.get('Valor (R$)') or 0) for c in compras)
        tct=sum(float(c.get('Valor (R$)',0) or 0) for c in contas)
        tf=sum(float(f.get('Valor Parcela (R$)') or f.get('Valor Total (R$)') or f.get('Valor (R$)') or 0) for f in filho)
        saldo=rec-des-tc-tct-tf
        cat_d={}
        for t in trans:
            if t.get('Tipo')=='Despesa': cat=t.get('Categoria','Outros'); cat_d[cat]=cat_d.get(cat,0)+float(t.get('Valor (R$)',0) or 0)
        cat_c={}
        for c in compras: cat=c.get('Categoria','Outros'); vp=float(c.get('Valor Parcela (R$)') or c.get('Valor Total (R$)') or 0); cat_c[cat]=cat_c.get(cat,0)+vp
        cat_f={}
        for f in filho:
            cat=f.get('Categoria','Outros')
            vf=float(f.get('Valor Parcela (R$)') or f.get('Valor Total (R$)') or f.get('Valor (R$)') or 0)
            cat_f[cat]=cat_f.get(cat,0)+vf

        monthly={}
        def _madd(mes, key, val):
            monthly.setdefault(mes,{'receitas':0,'despesas':0,'compras':0,'contas':0,'filho':0})
            monthly[mes][key]+=val

        def _safe_dia(raw):
            if raw is None: return 10
            try: return int(raw)
            except:
                try: return datetime.strptime(str(raw)[:10],'%Y-%m-%d').day
                except: return 10

        def _mes_compra(c):
            dv=str(c.get('Data Vencimento') or '')[:10]
            if len(dv)==10: return dv[:7]
            base=str(c.get('Data') or c.get('Data Compra') or '')[:10]
            if not base: return ''
            dia=_safe_dia(c.get('Dia Vencimento'))
            pa=int(c.get('Parcela Atual') or 1)
            try:
                b=datetime.strptime(base,'%Y-%m-%d')
                t2=b+relativedelta(months=pa-1)
                ld=cal_mod.monthrange(t2.year,t2.month)[1]
                return t2.replace(day=min(dia,ld)).strftime('%Y-%m')
            except: return base[:7]

        for t in trans:
            dt=str(t.get('Data','') or '')[:7]
            if not dt: continue
            val=float(t.get('Valor (R$)',0) or 0)
            _madd(dt,'receitas' if t.get('Tipo')=='Receita' else 'despesas',val)

        for c in compras:
            mes=_mes_compra(c)
            if not mes: continue
            vp=float(c.get('Valor Parcela (R$)') or c.get('Valor Total (R$)') or c.get('Valor (R$)') or 0)
            _madd(mes,'compras',vp)

        for c in contas:
            mr=str(c.get('Mes Referencia','') or '')
            if '/' in mr:
                p=mr.split('/'); mes=f"{p[1]}-{p[0].zfill(2)}"
            else: mes=datetime.now().strftime('%Y-%m')
            _madd(mes,'contas',float(c.get('Valor (R$)',0) or 0))

        for f in filho:
            dt=str(f.get('Data Vencimento') or f.get('Data','') or '')[:7]
            if not dt: continue
            vf=float(f.get('Valor Parcela (R$)') or f.get('Valor Total (R$)') or f.get('Valor (R$)') or 0)
            _madd(dt,'filho',vf)

        cs={'Pago':0,'Pendente':0,'Atrasado':0}
        for c in contas: s=c.get('Status','Pendente'); cs[s]=cs.get(s,0)+1
        return jsonify({'success':True,'resumo':{'receitas':rec,'despesas':des,'total_compras':tc,'total_contas':tct,'total_filho':tf,'saldo':saldo,'num_transacoes':len(trans),'num_compras':len(compras),'num_contas':len(contas),'num_filho':len(filho)},'categorias':cat_d,'categorias_compras':cat_c,'categorias_filho':cat_f,'evolucao_mensal':monthly,'contas_status':cs})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

# ── EVOLUÇÃO ──────────────────────────────────────────────────────────────────
@app.route('/api/evolucao',methods=['GET'])
def evolucao():
    try:
        inicio=request.args.get('inicio',''); fim=request.args.get('fim','')
        trans=load_sheet(SHEETS['transacoes']); compras=load_sheet(SHEETS['compras'])
        contas=load_sheet(SHEETS['contas']); filho=load_sheet(SHEETS['filho'])
        monthly=_build_projected_entries(trans,compras,contas,inicio,fim)
        for f in filho:
            dt=str(f.get('Data Vencimento') or f.get('Data','') or '')
            if not dt or len(dt)<7: continue
            if inicio and dt<inicio: continue
            if fim and dt>fim: continue
            mes=dt[:7]
            monthly.setdefault(mes,{'receitas':0,'despesas':0,'compras':0,'contas':0,'filho':0})
            vf=float(f.get('Valor Parcela (R$)') or f.get('Valor Total (R$)') or f.get('Valor (R$)') or 0)
            monthly[mes]['filho']+=vf
            rec_f = str(f.get('Recorrente') or 'Não')
            np_f  = int(f.get('Num Parcelas') or 1)
            if rec_f == 'Sim' and np_f <= 1:
                try: cur_f = datetime.strptime(mes + '-01', '%Y-%m-%d')
                except: continue
                horizon_ev = (datetime.today() + relativedelta(months=PROJECTION_MONTHS)).strftime('%Y-%m-%d')
                for _ in range(PROJECTION_MONTHS):
                    cur_f += relativedelta(months=1)
                    fm = cur_f.strftime('%Y-%m')
                    if cur_f.strftime('%Y-%m-%d') > horizon_ev: break
                    if fim and fm > fim[:7]: break
                    if inicio and fm < inicio[:7]: continue
                    monthly.setdefault(fm,{'receitas':0,'despesas':0,'compras':0,'contas':0,'filho':0})
                    monthly[fm]['filho'] += vf
        result=[]
        for mes in sorted(monthly.keys()):
            d=monthly[mes]; ts=d['despesas']+d['compras']+d['contas']+d['filho']
            result.append({'mes':mes,'receitas':round(d['receitas'],2),'despesas':round(d['despesas'],2),'compras':round(d['compras'],2),'contas':round(d['contas'],2),'filho':round(d['filho'],2),'total_saidas':round(ts,2),'saldo':round(d['receitas']-ts,2)})
        return jsonify({'success':True,'data':result})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

def _build_projected_entries(trans,compras,contas,inicio=None,fim=None):
    today=datetime.today(); horizon=(today+relativedelta(months=PROJECTION_MONTHS)).strftime('%Y-%m-%d')
    monthly={}
    def add(mes,key,val):
        monthly.setdefault(mes,{'receitas':0,'despesas':0,'compras':0,'contas':0,'filho':0}); monthly[mes][key]+=val
    def in_range(mk):
        if inicio and mk<inicio[:7]: return False
        if fim and mk>fim[:7]: return False
        return True
    for t in trans:
        dt=str(t.get('Data','') or '');
        if not dt or len(dt)<7: continue
        mes=dt[:7]; val=float(t.get('Valor (R$)',0) or 0); tipo=t.get('Tipo',''); rec=str(t.get('Recorrente','Não') or 'Não')
        if in_range(mes): add(mes,'receitas' if tipo=='Receita' else 'despesas',val)
        if tipo=='Receita' and rec=='Sim':
            cur=datetime.strptime(dt[:10],'%Y-%m-%d')
            for _ in range(PROJECTION_MONTHS):
                cur+=relativedelta(months=1); fm=cur.strftime('%Y-%m')
                if cur.strftime('%Y-%m-%d')>horizon: break
                if in_range(fm): add(fm,'receitas',val)
    for c in compras:
        dv=str(c.get('Data Vencimento') or c.get('Data Compra') or '')[:10]
        if not dv or len(dv)<7: continue
        mes=dv[:7]; vp=float(c.get('Valor Parcela (R$)') or c.get('Valor Total (R$)') or 0)
        if in_range(mes): add(mes,'compras',vp)
    for c in contas:
        rec=str(c.get('Recorrente','Sim') or 'Sim'); mr=str(c.get('Mes Referencia','') or ''); val=float(c.get('Valor (R$)',0) or 0)
        if '/' in mr: p=mr.split('/'); mes=f"{p[1]}-{p[0].zfill(2)}"
        else: mes=today.strftime('%Y-%m')
        if in_range(mes): add(mes,'contas',val)
        if rec=='Sim':
            try: cur=datetime.strptime(mes+'-01','%Y-%m-%d')
            except: continue
            for _ in range(PROJECTION_MONTHS):
                cur+=relativedelta(months=1); fm=cur.strftime('%Y-%m')
                if cur.strftime('%Y-%m-%d')>horizon: break
                if in_range(fm): add(fm,'contas',val)
    return monthly

@app.route('/api/download-excel')
def download(): return send_file(EXCEL_FILE,as_attachment=True,download_name='financas_casa.xlsx')

def _del(sheet,iid,comp_col=None):
    try:
        wb=openpyxl.load_workbook(EXCEL_FILE); ws=wb[sheet]
        for row in ws.iter_rows(min_row=2):
            if row[0].value==iid:
                if comp_col: del_file(ws.cell(row=row[0].row,column=comp_col).value)
                ws.delete_rows(row[0].row); break
        wb.save(EXCEL_FILE); _update_resumo(); return jsonify({'success':True})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

def _update_resumo():
    try:
        wb=openpyxl.load_workbook(EXCEL_FILE)
        ws_r=wb[SHEETS['resumo']]; ws_t=wb[SHEETS['transacoes']]
        ws_c=wb[SHEETS['compras']]; ws_ct=wb[SHEETS['contas']]; ws_f=wb[SHEETS['filho']]
        for row in ws_r.iter_rows(min_row=2):
            for cell in row: cell.value=None
        monthly={}
        def add(mes,key,val):
            monthly.setdefault(mes,{'receitas':0,'despesas':0,'compras':0,'contas':0,'filho':0}); monthly[mes][key]+=val
        for row in ws_t.iter_rows(min_row=2,values_only=True):
            if not row[0]: continue
            if row[4]=='Receita': add(str(row[1] or '')[:7],'receitas',float(row[6] or 0))
            else: add(str(row[1] or '')[:7],'despesas',float(row[6] or 0))
        hdrs_c=[ws_c.cell(1,c).value for c in range(1,ws_c.max_column+1)]
        idx_vp=hdrs_c.index('Valor Parcela (R$)') if 'Valor Parcela (R$)' in hdrs_c else 7
        idx_dv=hdrs_c.index('Data Vencimento') if 'Data Vencimento' in hdrs_c else 13
        for row in ws_c.iter_rows(min_row=2,values_only=True):
            if not row[0]: continue
            dt=str(row[idx_dv] or row[1] or '')[:7]; vp=float(row[idx_vp] or row[6] or 0); add(dt,'compras',vp)
        for row in ws_ct.iter_rows(min_row=2,values_only=True):
            if not row[0]: continue
            mr=str(row[7] or ''); dt=f"{mr.split('/')[1]}-{mr.split('/')[0].zfill(2)}" if '/' in mr else datetime.now().strftime('%Y-%m')
            add(dt,'contas',float(row[3] or 0))
        hdrs_f=[ws_f.cell(1,c).value for c in range(1,ws_f.max_column+1)]
        idx_fvp = hdrs_f.index('Valor Parcela (R$)') if 'Valor Parcela (R$)' in hdrs_f else (hdrs_f.index('Valor (R$)') if 'Valor (R$)' in hdrs_f else 5)
        idx_fdv = hdrs_f.index('Data Vencimento') if 'Data Vencimento' in hdrs_f else (hdrs_f.index('Data') if 'Data' in hdrs_f else 1)
        for row in ws_f.iter_rows(min_row=2,values_only=True):
            if not row[0]: continue
            dt=str(row[idx_fdv] or row[1] or '')[:7]
            vf=float(row[idx_fvp] or row[5] or 0)
            add(dt,'filho',vf)
        for idx,mes in enumerate(sorted(monthly.keys()),2):
            d=monthly[mes]; ts=d['despesas']+d['compras']+d['contas']+d['filho']; s=d['receitas']-ts
            ws_r.append([mes,d['receitas'],ts,d['compras'],d['contas'],d['filho'],s,'OK' if s>=0 else 'NEG'])
            _style_row(ws_r,idx,alt=(idx%2==0))
            for col in [2,3,4,5,6,7]: ws_r.cell(row=idx,column=col).number_format='R$ #,##0.00'
        wb.save(EXCEL_FILE)
    except Exception as e: print(f"[resumo] {e}")

if __name__=='__main__':
    init_excel()
    app.run(debug=True,port=5000)

try:
    init_excel()
except Exception as _e:
    print(f"[init] {_e}")